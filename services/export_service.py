"""
Excel export for the admin panel.

Two exports, both generated fresh from the database on demand (so they
always reflect current bookings — there's no stale file to keep in sync):

  - build_participants_workbook: one sheet per activity, listing the
    guests booked on it plus an attendance column.
  - build_contacts_workbook: a single sheet of every registered guest's
    contact details.

Files are written to a temp path and returned for sending via the bot.
openpyxl is a runtime dependency (see requirements.txt).
"""

import re
from datetime import datetime
from pathlib import Path
from tempfile import gettempdir

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from db.crud import admin as admin_crud
from db.models import Activity, BookingStatus
from utils.time_utils import format_time, format_time_range

_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
_HEADER_FILL = PatternFill("solid", start_color="365F91")
_BODY_FONT = Font(name="Arial")
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _style_header(sheet, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER


def _autosize(sheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width


def _safe_sheet_title(raw: str, used: set[str]) -> str:
    """Excel sheet titles: max 31 chars, no []:*?/\\, and must be unique."""
    title = re.sub(r"[\[\]:*?/\\]", " ", raw).strip()[:31] or "Sheet"
    base, n = title, 1
    while title in used:
        suffix = f" ({n})"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(title)
    return title


def _status_label(status: BookingStatus) -> str:
    return {
        BookingStatus.CONFIRMED: "Підтверджено",
        BookingStatus.PENDING_CONFIRMATION: "Очікує підтвердження",
        BookingStatus.ATTENDED: "Був присутній",
    }.get(status, status.value)


async def build_participants_workbook(session: AsyncSession) -> Path:
    """
    One sheet per regular activity, plus a single combined sheet for all
    consultation slots (not one per 15-min slot).
    """
    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()

    activities = await admin_crud.all_activities(session)

    consult_activities = [a for a in activities if a.is_consultation_slot]
    regular_activities = [a for a in activities if not a.is_consultation_slot]

    for activity in regular_activities:
        participants = await admin_crud.participants_for_activity(session, activity.id)
        title = _safe_sheet_title(f"Д{activity.day} {activity.title}", used_titles)
        sheet = wb.create_sheet(title)

        sheet["A1"] = f"{activity.title} — {format_time_range(activity.start_time, activity.end_time)}"
        sheet["A1"].font = Font(bold=True, name="Arial", size=12)
        sheet["A2"] = f"Зайнято: {len(participants)} з {activity.capacity}"
        sheet["A2"].font = _BODY_FONT

        headers = ["№", "Прізвище та ім'я", "Телефон", "Email", "Статус", "Присутність"]
        sheet.append([])
        sheet.append(headers)
        header_row = sheet.max_row
        for col in range(1, len(headers) + 1):
            c = sheet.cell(row=header_row, column=col)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = _CENTER

        for i, (booking, user) in enumerate(participants, start=1):
            attended = "Так" if booking.status == BookingStatus.ATTENDED else ""
            sheet.append([i, user.full_name, user.phone, user.email, _status_label(booking.status), attended])

        _autosize(sheet, [5, 28, 16, 28, 22, 14])

    # Combined consultations sheet.
    if consult_activities:
        sheet = wb.create_sheet(_safe_sheet_title("Консультації", used_titles))
        first, last = consult_activities[0], consult_activities[-1]
        present = 0

        sheet["A1"] = f"Консультації Анни Баринової — {format_time_range(first.start_time, last.end_time)}"
        sheet["A1"].font = Font(bold=True, name="Arial", size=12)
        sheet.append([])
        sheet.append([])  # placeholder for occupancy (filled after counting)
        headers = ["Час", "Прізвище та ім'я", "Телефон", "Email", "Статус", "Присутність"]
        sheet.append(headers)
        header_row = sheet.max_row
        for col in range(1, len(headers) + 1):
            c = sheet.cell(row=header_row, column=col)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = _CENTER

        for activity in consult_activities:
            participants = await admin_crud.participants_for_activity(session, activity.id)
            for booking, user in participants:
                present += 1
                attended = "Так" if booking.status == BookingStatus.ATTENDED else ""
                sheet.append([
                    format_time(activity.start_time), user.full_name, user.phone,
                    user.email, _status_label(booking.status), attended,
                ])

        sheet["A3"] = f"Зайнято слотів: {present} з {len(consult_activities)}"
        sheet["A3"].font = _BODY_FONT
        _autosize(sheet, [10, 28, 16, 28, 22, 14])

    if not wb.sheetnames:
        wb.create_sheet("Порожньо")

    return _save(wb, "participants")


async def build_contacts_workbook(session: AsyncSession) -> Path:
    """A single sheet of all registered guests' contact details."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Контакти"

    headers = ["№", "Прізвище та ім'я", "Телефон", "Email", "Зареєстровано"]
    sheet.append(headers)
    _style_header(sheet, len(headers))

    guests = await admin_crud.all_guests(session)
    for i, user in enumerate(guests, start=1):
        registered = user.created_at.strftime("%Y-%m-%d %H:%M") if user.created_at else ""
        sheet.append([i, user.full_name, user.phone, user.email, registered])
        for col in range(1, len(headers) + 1):
            sheet.cell(row=i + 1, column=col).font = _BODY_FONT

    _autosize(sheet, [5, 28, 16, 28, 18])
    return _save(wb, "contacts")


def _save(wb: Workbook, prefix: str) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = Path(gettempdir()) / f"wellness_{prefix}_{stamp}.xlsx"
    wb.save(path)
    return path
